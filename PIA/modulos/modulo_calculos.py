import statistics as st
import openpyxl as xl

def media_duracion(archivo):
    wb = xl.load_workbook(archivo)
    ws = wb.active

    datos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos.append(fila[4])

    m = st.mean(datos)

    ws["H1"] = "Media de duracion de cancion"
    ws["H2"] = m
    wb.save(archivo)

    return m


def varianza_discos(archivo):
    wb = xl.load_workbook(archivo)
    ws = wb.active

    datos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos.append(fila[4])

    v = st.variance(datos)

    ws["I1"] = "Varianza de duracion de cancion"
    ws["I2"] = v
    wb.save(archivo)

    return v


def porcentaje_duracion(archivo):
    wb = xl.load_workbook(archivo)
    ws = wb.active

    nombres = []
    datos = []

    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombres.append(fila[0])
        datos.append(fila[4])

    total = sum(datos)

    i = 2
    for d in datos:
        if total != 0:
            p = d * 100 / total
        else:
            p = 0

        ws["F1"] = "Porcentaje de discografia"
        ws["F" + str(i)] = p
        print("\t",nombres[i-2], "es el",p, "% de la discografia")

        i += 1

    wb.save(archivo)

def media_pistas(archivo):
    
    wb = xl.load_workbook(archivo)
    ws = wb.active

    datos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        datos.append(fila[3])

    m = st.mean(datos)

    ws["J1"] = "Media de total de pistas"
    ws["J2"] = m
    wb.save(archivo)

    return m
